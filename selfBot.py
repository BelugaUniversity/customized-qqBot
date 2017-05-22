# -*- coding: UTF-8 -*-
from qqbot import QQBotSlot as qqbotslot, RunBot
import openpyxl
import time
@qqbotslot
def onQQMessage(bot, contact, member, content):
    array = content.split('_')
    if content == '-hello':
        bot.SendTo(contact,  '@' + member.name + "你好，我是QQ机器人")
    if content == '-stop':
        bot.SendTo(contact, "QQ机器人已关闭")
        bot.Stop()
    if '@点点书童' in content or '@ME' in content:
        print content
        if '_学习打卡_' in content:
            wb = openpyxl.load_workbook('record.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            maxRow = sheet.max_row
            maxColumn = sheet.max_column  
            for j in range(0, len(array)):
                if '@点点书童' or '@ME' in content in array[j]:
                    array[j] = array[j].replace('@点点书童', '')
                    array[j] = array[j].replace('[@ME] ', '')
                    
            #判断是否是当天时间
            if array[0] != time.strftime("%Y%m%d"):
                bot.SendTo(contact,  '@' + member.name + " 童鞋你可能是穿越了~")
            # 循环判断有无相同日期和姓名self.
            elif not array[1] in member.name and not "胡蒙" in member.name:
                bot.SendTo(contact,  '@' + member.name + " 君子行不改名坐不改姓，重新报上名来！")
            else:
                for i in range(2, maxColumn + 1):
                    if int(array[0]) == sheet.cell(row = 2, column = i).value:    #日期相同
                        for j in range(2, maxRow + 1):
                            if  array[1].decode('utf8').encode('gb2312') == sheet.cell(row = j, column = 1).value:  #名字相同
                                sheet.cell(row = j, column = i).value = array[3]
                                print "1"
                                break
                            elif j == maxRow and array[1].decode('utf8').encode('gb2312') != sheet.cell(row = j, column = 1).value:                                 #没找到相同名字，直接创建
                                sheet.cell(row = j + 1, column = 1).value = array[1]
                                sheet.cell(row = j + 1, column = i).value = array[3]
                                print "2"
                                break
                        
                        break
                    elif i == maxColumn and int(array[0]) != sheet.cell(row=2, column=i).value:                                          #没找到日期创建
                        sheet.cell(row = 2, column = i + 1).value = int(array[0])
                        for j in range(2, maxRow + 1):
                            if array[1].decode('utf8').encode('gb2312') == sheet.cell(row = j, column = 1).value:  #名字相同
                                sheet.cell(row = j, column = i + 1).value = array[3]
                                print "3"
                                break
                            elif j == maxRow and array[1].decode('utf8').encode('gb2312') != sheet.cell(row = j, column = 2).value:                                 #没找到相同名字，直接创建
                                sheet.cell(row = j + 1, column = 1).value = array[1]
                                sheet.cell(row = j + 1, column=i + 1).value = array[3]
                                print "4"
                                break
                        break

                wb.save('record.xlsx')
                bot.SendTo(contact,  '@' + member.name + " 你说的一切已成为呈堂证供！")
        else:
            bot.SendTo(contact,  '@' + member.name + " 每日打卡，好好学习，天天向上！")
RunBot()
